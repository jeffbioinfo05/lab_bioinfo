"""
Coleta métricas do pipeline assembly_prokaryote e gera Excel.
Chamado via Snakemake: snakemake.input, snakemake.output, snakemake.params
"""
import json, re, os, sys
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils.dataframe import dataframe_to_rows

rows = []
for fastp_json, quast_tsv, busco_txt, checkm2_tsv, gff_file in zip(
    snakemake.input.fastp, snakemake.input.quast,
    snakemake.input.busco, snakemake.input.checkm2,
    snakemake.input.gff):

    sample = os.path.basename(os.path.dirname(gff_file))
    r = {"sample": sample}

    # fastp
    with open(fastp_json) as f:
        fp = json.load(f)
    r["raw_reads"]       = fp["summary"]["before_filtering"]["total_reads"]
    r["trimmed_reads"]   = fp["summary"]["after_filtering"]["total_reads"]
    r["q30_rate"]        = fp["summary"]["after_filtering"]["q30_rate"]
    r["pct_adapter"]     = round(fp.get("adapter_cutting", {}).get("adapter_trimmed_reads", 0)
                                  / r["raw_reads"] * 100, 2)

    # QUAST
    try:
        qdf = pd.read_csv(quast_tsv, sep="\t", header=None, names=["metric","value"])
        qmap = dict(zip(qdf.metric, qdf.value))
        r["n50"]            = qmap.get("N50", "")
        r["num_contigs"]    = qmap.get("# contigs", "")
        r["total_length_mb"]= round(int(qmap.get("Total length", 0)) / 1e6, 2)
        r["largest_contig"] = qmap.get("Largest contig", "")
        r["gc_percent"]     = qmap.get("GC (%)", "")
    except Exception: pass

    # BUSCO
    try:
        with open(busco_txt) as f:
            txt = f.read()
        m = re.search(r"C:(\S+)%\[S:(\S+)%,D:(\S+)%\],F:(\S+)%,M:(\S+)%", txt)
        if m:
            r["busco_complete"] = float(m.group(1))
            r["busco_duplicated"]= float(m.group(3))
            r["busco_fragmented"]= float(m.group(4))
            r["busco_missing"]   = float(m.group(5))
    except Exception: pass

    # CheckM2
    try:
        cdf = pd.read_csv(checkm2_tsv, sep="\t")
        r["checkm2_completeness"]  = cdf["Completeness"].iloc[0]
        r["checkm2_contamination"] = cdf["Contamination"].iloc[0]
    except Exception: pass

    # GFF annotation count
    try:
        with open(gff_file) as f:
            lines = [l for l in f if not l.startswith("#") and "\tCDS\t" in l]
        r["num_cds"] = len(lines)
    except Exception: pass

    rows.append(r)

df = pd.DataFrame(rows)

wb = Workbook()
ws = wb.active
ws.title = "Assembly QC"

header_fill = PatternFill("solid", fgColor="2E86AB")
header_font = Font(color="FFFFFF", bold=True)

for r_idx, row in enumerate(dataframe_to_rows(df, index=False, header=True), 1):
    ws.append(row)
    if r_idx == 1:
        for cell in ws[r_idx]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center")

for col in ws.columns:
    ws.column_dimensions[col[0].column_letter].width = max(len(str(c.value or "")) for c in col) + 2

wb.save(snakemake.output[0])
print(f"Planilha gerada: {snakemake.output[0]}", file=sys.stderr)
